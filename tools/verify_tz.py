"""Сверка готового отчёта с требованиями ТЗ. Ничего не меняет, только проверяет."""
from __future__ import annotations

import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
files = [p for p in (ROOT / "out").glob("Отчет_котельные_*.xlsx") if "ДЕМО" not in p.name]
files.sort(key=lambda p: p.stat().st_mtime)
if not files:
    print("Нет отчётов в out/")
    raise SystemExit(1)

path = files[-1]
print(f"Файл: {path.name}\n")
wb = load_workbook(path)
ws = wb[wb.sheetnames[0]]
print(f"Листы: {wb.sheetnames}")

# 1. Колонки по ТЗ
header_row = None
for r in range(1, 15):
    if str(ws.cell(r, 1).value or "").strip().lower().startswith("наименование"):
        header_row = r
        break
print(f"\n1) Строка заголовков: {header_row}")
headers = [ws.cell(header_row, c).value for c in range(1, 9)]
for i, h in enumerate(headers, 1):
    print(f"   {i}. {h}")

required = [
    "Наименование обьекта",
    "Потребление Тепла, Гкал",
    "Расход газа Нм3/ч",
    "Затраты по котельной",
    "КПД Котельной %",
]
print("\n   Требуемые по ТЗ:")
for i, req in enumerate(required):
    got = str(headers[i] or "")
    mark = "совпадает" if req.lower().replace(" ", "") in got.lower().replace(" ", "") else f"ОТЛИЧАЕТСЯ -> «{got}»"
    print(f"   {i+1}. {req}: {mark}")

# 2. Подсветка КПД ниже 80
print("\n2) Условное форматирование (красный при КПД < 80):")
found = False
for rng in ws.conditional_formatting:
    for rule in rng.rules:
        print(f"   диапазон {rng.sqref}: {rule.operator} {rule.formula}")
        found = True
if not found:
    print("   НЕ НАЙДЕНО")

# 3. КПД формулой
print("\n3) КПД в ячейках:")
sample = ws.cell(header_row + 1, 5).value
print(f"   {str(sample)[:110]}")
print(f"   формула: {'да' if str(sample).startswith('=') else 'НЕТ'}")

# 4. Значения считаем сами из Гкал и Нм3: openpyxl формулы не вычисляет
print("\n4) Проверка значений (расчёт из колонок B и C):")
vals = []
for r in range(header_row + 1, ws.max_row + 1):
    heat, gas = ws.cell(r, 2).value, ws.cell(r, 3).value
    name = ws.cell(r, 1).value
    if isinstance(heat, (int, float)) and isinstance(gas, (int, float)) and gas:
        vals.append((name, heat * 1_000_000 / (gas * 8200) * 100))
print(f"   строк с теплом и газом: {len(vals)}")
if vals:
    lo = [x for x in vals if x[1] < 80]
    hi = [x for x in vals if x[1] > 100]
    print(f"   ниже 80% — подсвечиваются красным: {len(lo)}")
    for n, v in lo:
        print(f"      {v:5.1f}%  {str(n)[:52]}")
    print(f"   выше 100% — КПД скрыт формулой: {len(hi)}")
    for n, v in hi:
        print(f"      {v:5.1f}%  {str(n)[:52]}")
    ok = [x for x in vals if 80 <= x[1] <= 100]
    print(f"   в норме 80-100%: {len(ok)}")
