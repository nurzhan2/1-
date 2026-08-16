"""Получение затрат по объектам из 1С.

По ТЗ 1С опрашивается раз в месяц: нужны «Затраты по котельной» в разрезе объектов
из переданного заказчиком перечня.

Пока доступ к базе не выдан, работает режим ``manual``: затраты берутся из файла
``data/costs.xlsx`` (две колонки — «Объект», «Затраты»). Это позволяет собирать
полный отчёт и согласовать его форму, не дожидаясь интеграции. Переключение на
боевой режим — одна строка в .env, код отчёта при этом не меняется.
"""
from __future__ import annotations

import logging
import re
from datetime import date
from pathlib import Path
from typing import Optional

import httpx

from .config import OneCSettings

log = logging.getLogger(__name__)


class OneCError(RuntimeError):
    pass


class OneCClient:
    def __init__(self, cfg: OneCSettings):
        self.cfg = cfg

    async def fetch_costs(self, start: date, end: date) -> dict[str, float]:
        """Возвращает {наименование объекта: сумма затрат} за период."""
        mode = self.cfg.mode
        if mode == "odata":
            return await self._fetch_odata(start, end)
        if mode == "http":
            return await self._fetch_http(start, end)
        if mode == "manual":
            return self._fetch_manual()
        raise OneCError(f"Неизвестный режим ONEC_MODE={mode!r}")

    # -------------------------------------------------------------------- OData
    async def _fetch_odata(self, start: date, end: date) -> dict[str, float]:
        if not self.cfg.base_url:
            raise OneCError("Не задан ONEC_BASE_URL для режима odata")
        url = (
            f"{self.cfg.base_url.rstrip('/')}/odata/standard.odata/"
            f"{self.cfg.odata_resource}(StartPeriod=datetime'{start:%Y-%m-%dT00:00:00}',"
            f"EndPeriod=datetime'{end:%Y-%m-%dT23:59:59}')"
        )
        params = {"$format": "json"}
        auth = (self.cfg.username, self.cfg.password) if self.cfg.username else None

        async with httpx.AsyncClient(timeout=self.cfg.timeout, auth=auth) as client:
            resp = await client.get(url, params=params)
        if resp.status_code == 401:
            raise OneCError("1С отклонила авторизацию (401). Проверьте логин/пароль и права на OData.")
        if resp.status_code >= 400:
            raise OneCError(f"1С вернула {resp.status_code}: {resp.text[:300]}")

        payload = resp.json()
        rows = payload.get("value", payload if isinstance(payload, list) else [])
        return self._aggregate(rows, self.cfg.field_object, self.cfg.field_amount)

    # -------------------------------------------------------- самописный HTTP-сервис
    async def _fetch_http(self, start: date, end: date) -> dict[str, float]:
        if not self.cfg.base_url:
            raise OneCError("Не задан ONEC_BASE_URL для режима http")
        url = self.cfg.base_url.rstrip("/") + self.cfg.http_path
        auth = (self.cfg.username, self.cfg.password) if self.cfg.username else None
        params = {"start": start.isoformat(), "end": end.isoformat()}

        async with httpx.AsyncClient(timeout=self.cfg.timeout, auth=auth) as client:
            resp = await client.get(url, params=params)
        if resp.status_code >= 400:
            raise OneCError(f"HTTP-сервис 1С вернул {resp.status_code}: {resp.text[:300]}")

        payload = resp.json()
        if isinstance(payload, dict) and "items" in payload:
            payload = payload["items"]
        if isinstance(payload, dict):
            return {str(k): _to_float(v) or 0.0 for k, v in payload.items()}
        return self._aggregate(payload, self.cfg.field_object, self.cfg.field_amount)

    # ------------------------------------------------------------------- вручную
    def _fetch_manual(self) -> dict[str, float]:
        path = Path(self.cfg.manual_file)
        if not path.exists():
            log.warning("Файл затрат %s не найден — колонка «Затраты» останется пустой", path)
            return {}

        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True, read_only=True)
        result: dict[str, float] = {}
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells = [c for c in row if c is not None]
                if len(cells) < 2:
                    continue
                name = str(cells[0]).strip()
                amount = _to_float(cells[1])
                if not name or amount is None:
                    continue
                if "объект" in name.lower() or "наименование" in name.lower():
                    continue
                result[name] = result.get(name, 0.0) + amount
        wb.close()
        log.info("Затраты из файла %s: %d объектов", path.name, len(result))
        return result

    # -------------------------------------------------------------------- helpers
    @staticmethod
    def _aggregate(rows, field_object: str, field_amount: str) -> dict[str, float]:
        result: dict[str, float] = {}
        for row in rows or []:
            if not isinstance(row, dict):
                continue
            name = row.get(field_object) or row.get(f"{field_object}_Key") or ""
            if isinstance(name, dict):
                name = name.get("Description") or name.get("Наименование") or ""
            name = str(name).strip()
            amount = _to_float(row.get(field_amount))
            if not name or amount is None:
                continue
            result[name] = result.get(name, 0.0) + amount
        return result


def load_cost_mapping(path) -> dict[str, int]:
    """Читает таблицу соответствий «название 1С → номер котельной ЛЭРС».

    Ожидает файл с двумя колонками: название статьи затрат в 1С и номер котельной
    из перечня. Возвращает {нормализованное имя 1С: номер}. Если файла нет —
    пустой словарь (тогда используется сопоставление по названиям).
    """
    p = Path(path)
    if not p.exists():
        return {}
    from openpyxl import load_workbook

    wb = load_workbook(p, data_only=True, read_only=True)
    mapping: dict[str, int] = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = [c for c in row if c is not None]
            if len(cells) < 2:
                continue
            name = str(cells[0]).strip()
            number = _to_int(cells[1])
            if not name or number is None:
                continue
            low = name.lower()
            if "1с" in low or "назван" in low or "котельн" in low:  # строка-заголовок
                continue
            mapping[_normalize(name)] = number
    wb.close()
    log.info("Таблица соответствий 1С↔ЛЭРС: %d строк из %s", len(mapping), p.name)
    return mapping


def costs_by_number(costs: dict[str, float], mapping: dict[str, int]) -> dict[int, float]:
    """Раскладывает затраты 1С по номерам котельных через таблицу соответствий."""
    out: dict[int, float] = {}
    if not mapping:
        return out
    for name, amount in costs.items():
        number = mapping.get(_normalize(name))
        if number is not None:
            out[number] = out.get(number, 0.0) + amount
    return out


def _to_int(value) -> Optional[int]:
    f = _to_float(value)
    return int(f) if f is not None else None


def _to_float(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    text = re.sub(r"[^\d.\-]", "", text)
    if not text or text in {"-", "."}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def match_costs(costs: dict[str, float], title: str) -> Optional[float]:
    """Сопоставление названия объекта из ЛЭРС с названием статьи затрат в 1С.

    Названия в двух системах совпадают редко, поэтому сравниваем нормализованные
    строки, а затем ищем вхождение по значимым словам. Все неоднозначные случаи
    логируются, чтобы заказчик мог один раз согласовать таблицу соответствий.
    """
    if not costs:
        return None
    norm_title = _normalize(title)
    if norm_title in {_normalize(k) for k in costs}:
        for key, value in costs.items():
            if _normalize(key) == norm_title:
                return value

    title_words = {w for w in norm_title.split() if len(w) > 3}
    best_key, best_score = None, 0
    for key, value in costs.items():
        key_words = {w for w in _normalize(key).split() if len(w) > 3}
        if not key_words:
            continue
        score = len(title_words & key_words)
        if score > best_score:
            best_key, best_score = key, score
    if best_key is not None and best_score >= 2:
        log.debug("Затраты: «%s» сопоставлено с «%s»", title, best_key)
        return costs[best_key]
    return None


def _normalize(text: str) -> str:
    text = text.lower().replace("ё", "е")
    text = re.sub(r"[«»\"'(),.№\-_/]", " ", text)
    text = re.sub(r"\b(г|с|д|п|пос|птг|ст|ул|мбоу|гбуз|гапоу|мау|до|ооо)\b", " ", text)
    return " ".join(text.split())
