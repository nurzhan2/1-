"""Клиент 1С в трёх режимах: odata / http / manual."""
from __future__ import annotations

import re
from datetime import date

import httpx
import pytest
import respx
from openpyxl import Workbook

from app.onec_client import OneCClient, OneCError

from .conftest import load_fixture, make_onec_settings

START = date(2026, 7, 1)
END = date(2026, 7, 31)


async def test_manual_mode_reads_xlsx(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Объект", "Затраты"])
    ws.append(["Котельная A", "125 000,50"])
    ws.append(["Котельная B", 240000])
    ws.append([None, None])
    path = tmp_path / "costs.xlsx"
    wb.save(path)

    client = OneCClient(make_onec_settings(mode="manual", manual_file=path))
    costs = await client.fetch_costs(START, END)
    assert costs == {"Котельная A": 125000.5, "Котельная B": 240000.0}


async def test_manual_mode_missing_file_returns_empty(tmp_path):
    client = OneCClient(make_onec_settings(mode="manual", manual_file=tmp_path / "нет.xlsx"))
    assert await client.fetch_costs(START, END) == {}


async def test_odata_mode(monkeypatch):
    settings = make_onec_settings(mode="odata")
    with respx.mock(assert_all_called=True) as mock:
        mock.get(url__regex=r"http://onec\.test/odata/.*").mock(
            return_value=httpx.Response(200, json=load_fixture("onec_odata.json"))
        )
        costs = await OneCClient(settings).fetch_costs(START, END)
    assert costs == {
        "Котельная на ул. Строителей": 125000.5,
        "Котельная центральная": 240000.0,
    }


async def test_odata_unauthorized_raises():
    settings = make_onec_settings(mode="odata")
    with respx.mock(assert_all_called=True) as mock:
        mock.get(url__regex=r"http://onec\.test/odata/.*").mock(
            return_value=httpx.Response(401, text="Unauthorized")
        )
        with pytest.raises(OneCError):
            await OneCClient(settings).fetch_costs(START, END)


async def test_http_mode():
    settings = make_onec_settings(mode="http")
    with respx.mock(assert_all_called=True) as mock:
        mock.get(url__regex=r"http://onec\.test/hs/costs/period.*").mock(
            return_value=httpx.Response(200, json=load_fixture("onec_http.json"))
        )
        costs = await OneCClient(settings).fetch_costs(START, END)
    assert costs == {
        "Котельная на ул. Строителей": 99000.0,
        "Котельная южная": 51000.25,
    }


def test_cost_mapping_by_number(tmp_path):
    from app.onec_client import costs_by_number, load_cost_mapping

    wb = Workbook()
    ws = wb.active
    ws.append(["Название в 1С", "Номер котельной ЛЭРС"])  # заголовок пропускается
    ws.append(["Серково ДК", 12])
    ws.append(["Камешково ЦРБ", 21])
    path = tmp_path / "map.xlsx"
    wb.save(path)

    mapping = load_cost_mapping(path)
    assert len(mapping) == 2
    by = costs_by_number({"серково дк": 100.0, "Камешково ЦРБ": 50.0, "Прочее": 9.0}, mapping)
    assert by == {12: 100.0, 21: 50.0}  # «Прочее» без соответствия не попало


def test_build_report_prefers_mapping_over_name(tmp_path):
    from app.models import MeasurePoint
    from app.onec_client import load_cost_mapping
    from app.report import build_report

    wb = Workbook()
    ws = wb.active
    ws.append(["1С", "Номер"])
    ws.append(["Серково ДК", 12])
    path = tmp_path / "map.xlsx"
    wb.save(path)
    mapping = load_cost_mapping(path)

    points = [MeasurePoint(number=12, title="с. Серково, Дом культуры")]
    costs = {"Серково ДК": 13529.45}  # по названию бы не совпало, а по номеру — да
    report = build_report(points, {}, costs, START, END, cost_mapping=mapping)
    assert report.rows[0].costs_rub == 13529.45


async def test_unknown_mode_raises():
    settings = make_onec_settings(mode="сломанный")
    with pytest.raises(OneCError):
        await OneCClient(settings).fetch_costs(START, END)
