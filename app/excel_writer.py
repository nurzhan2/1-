"""Выгрузка отчёта в Excel по форме из ТЗ."""
from __future__ import annotations

import logging
from datetime import date
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName

from .models import DailyRecord, ReportData

log = logging.getLogger(__name__)

FONT = "Arial"

# Именованные диапазоны: КПД и подсветка считаются через них, а не через $B$3/$B$4.
# Так формулы читаются осмысленно и не «поедут» при вставке строк.
NAME_CALORIFIC = "ТеплотвГаза"
NAME_THRESHOLD = "ПорогКПД"

HEADERS = [
    "Наименование объекта",
    "Потребление тепла, Гкал",
    "Расход газа, Нм3",
    "Затраты по котельной, руб.",
    "КПД котельной, %",
    "№ точки в ЛЭРС",
    "Полнота данных",
    "Примечание",
]

WIDTHS = [58, 22, 18, 22, 18, 15, 16, 38]

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
ASSUMPTION_FILL = PatternFill("solid", fgColor="FFF2CC")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")
BAD_FONT = Font(name=FONT, size=10, color="9C0006", bold=True)
INPUT_FONT = Font(name=FONT, size=10, color="0000FF", bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Номера строк служебного блока
ROW_TITLE = 1
ROW_PERIOD = 2
ROW_CALORIFIC = 3
ROW_THRESHOLD = 4
ROW_HEADER = 6
ROW_FIRST_DATA = 7


def write_report(
    report: ReportData,
    output_path: Path,
    calorific_kcal: float,
    threshold: float,
    daily: Optional[dict[int, list[DailyRecord]]] = None,
    banner: Optional[str] = None,
) -> Path:
    """Создаёт .xlsx и возвращает путь к нему.

    ``banner`` — предупреждающая строка над таблицей (например, пометка о том,
    что файл сформирован на демонстрационных данных).
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    _write_head(ws, report, calorific_kcal, threshold)
    _define_names(wb, ws)
    if banner:
        cell = ws.cell(ROW_HEADER - 1, 1, banner)
        cell.font = Font(name=FONT, size=11, bold=True, color="9C0006")
        cell.fill = BAD_FILL
        ws.merge_cells(
            start_row=ROW_HEADER - 1, start_column=1, end_row=ROW_HEADER - 1, end_column=len(HEADERS)
        )
        ws.row_dimensions[ROW_HEADER - 1].height = 22
    last_row = _write_table(ws, report)
    _write_totals(ws, last_row)
    _apply_conditional_formatting(ws, last_row)
    _finalize_sheet(ws)

    if daily:
        _write_daily_sheet(wb, daily, report)
    _write_notes_sheet(wb, report, calorific_kcal, threshold)

    wb.save(output_path)
    log.info("Отчёт сохранён: %s", output_path)
    return output_path


def _define_names(wb, ws) -> None:
    """Заводит имена ТеплотвГаза (B3) и ПорогКПД (B4) для формул и подсветки."""
    sheet = quote_sheetname(ws.title)
    for name, cell in ((NAME_CALORIFIC, "$B$3"), (NAME_THRESHOLD, "$B$4")):
        if name in wb.defined_names:
            del wb.defined_names[name]
        wb.defined_names.add(DefinedName(name, attr_text=f"{sheet}!{cell}"))


# ------------------------------------------------------------------ шапка листа
def _write_head(ws, report: ReportData, calorific_kcal: float, threshold: float) -> None:
    ws.cell(ROW_TITLE, 1, "Сводный отчёт по котельным").font = Font(name=FONT, size=14, bold=True)
    ws.cell(ROW_PERIOD, 1, "Период:").font = Font(name=FONT, size=10, bold=True)
    ws.cell(ROW_PERIOD, 2, report.period_label).font = Font(name=FONT, size=10)

    ws.cell(ROW_CALORIFIC, 1, "Теплотворная способность газа, ккал/Нм3:").font = Font(
        name=FONT, size=10, bold=True
    )
    cell = ws.cell(ROW_CALORIFIC, 2, calorific_kcal)
    cell.font = INPUT_FONT
    cell.fill = ASSUMPTION_FILL
    cell.number_format = "#,##0"
    ws.cell(
        ROW_CALORIFIC, 3, "значение из ТЗ; изменить здесь — КПД пересчитается автоматически"
    ).font = Font(name=FONT, size=9, italic=True, color="808080")

    ws.cell(ROW_THRESHOLD, 1, "Порог КПД для подсветки, %:").font = Font(name=FONT, size=10, bold=True)
    cell = ws.cell(ROW_THRESHOLD, 2, threshold)
    cell.font = INPUT_FONT
    cell.fill = ASSUMPTION_FILL
    cell.number_format = "0"
    ws.cell(ROW_THRESHOLD, 3, "ниже порога ячейка КПД подсвечивается красным").font = Font(
        name=FONT, size=9, italic=True, color="808080"
    )


# ------------------------------------------------------------------- таблица
def _write_table(ws, report: ReportData) -> int:
    for col, (header, width) in enumerate(zip(HEADERS, WIDTHS), start=1):
        cell = ws.cell(ROW_HEADER, col, header)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[ROW_HEADER].height = 32

    row_idx = ROW_FIRST_DATA
    for row in report.rows:
        ws.cell(row_idx, 1, row.title).font = Font(name=FONT, size=10)
        ws.cell(row_idx, 1).alignment = Alignment(vertical="center", wrap_text=True)

        heat = ws.cell(row_idx, 2, row.heat_gcal)
        heat.number_format = "#,##0.000;-#,##0.000;-"

        gas = ws.cell(row_idx, 3, row.gas_nm3)
        gas.number_format = "#,##0;-#,##0;-"

        cost = ws.cell(row_idx, 4, row.costs_rub)
        cost.number_format = "#,##0.00 ₽;-#,##0.00 ₽;-"

        # КПД считается формулой, а не в Python: лист остаётся живым документом
        eff = ws.cell(
            row_idx,
            5,
            f'=IFERROR(B{row_idx}/(C{row_idx}*{NAME_CALORIFIC}/1000000)*100,"")',
        )
        eff.number_format = "0.0"

        ws.cell(row_idx, 6, row.point_number).number_format = "0"
        ws.cell(
            row_idx,
            7,
            f"{row.days_with_data}/{row.days_expected}" if row.days_expected else "",
        )
        ws.cell(row_idx, 8, row.comment).font = Font(name=FONT, size=9, color="808080")

        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row_idx, col)
            cell.border = BORDER
            if cell.font.name != FONT:
                cell.font = Font(name=FONT, size=10)
            if col in (2, 3, 4, 5, 6, 7):
                cell.alignment = Alignment(horizontal="center", vertical="center")
        row_idx += 1

    return row_idx - 1


def _write_totals(ws, last_row: int) -> None:
    if last_row < ROW_FIRST_DATA:
        return
    total_row = last_row + 1
    ws.cell(total_row, 1, "ИТОГО").font = Font(name=FONT, size=10, bold=True)
    for col in (2, 3, 4):
        letter = get_column_letter(col)
        cell = ws.cell(total_row, col, f"=SUM({letter}{ROW_FIRST_DATA}:{letter}{last_row})")
        cell.font = Font(name=FONT, size=10, bold=True)
        cell.number_format = ws.cell(ROW_FIRST_DATA, col).number_format
        cell.alignment = Alignment(horizontal="center")
    # В итоговом КПД учитываем только объекты, где есть учёт газа, иначе
    # тепло «безгазовых» точек завысило бы результат
    eff = ws.cell(
        total_row,
        5,
        f"=IFERROR(SUMPRODUCT((C{ROW_FIRST_DATA}:C{last_row}>0)*B{ROW_FIRST_DATA}:B{last_row})"
        f'/(C{total_row}*{NAME_CALORIFIC}/1000000)*100,"")',
    )
    eff.font = Font(name=FONT, size=10, bold=True)
    eff.number_format = "0.0"
    eff.alignment = Alignment(horizontal="center")
    for col in range(1, len(HEADERS) + 1):
        ws.cell(total_row, col).border = BORDER


def _apply_conditional_formatting(ws, last_row: int) -> None:
    if last_row < ROW_FIRST_DATA:
        return
    rng = f"E{ROW_FIRST_DATA}:E{last_row + 1}"
    ws.conditional_formatting.add(
        rng,
        CellIsRule(
            operator="lessThan",
            formula=[NAME_THRESHOLD],
            fill=BAD_FILL,
            font=BAD_FONT,
            stopIfTrue=False,
        ),
    )


def _finalize_sheet(ws) -> None:
    ws.freeze_panes = f"A{ROW_FIRST_DATA}"
    ws.auto_filter.ref = f"A{ROW_HEADER}:{get_column_letter(len(HEADERS))}{ws.max_row}"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.print_title_rows = f"{ROW_HEADER}:{ROW_HEADER}"
    # чтобы при печати/экспорте в PDF все колонки помещались в ширину листа
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


# ------------------------------------------------------- лист суточных данных
def _write_daily_sheet(wb, daily: dict[int, list[DailyRecord]], report: ReportData) -> None:
    ws = wb.create_sheet("Суточные данные")
    headers = ["№ точки", "Дата", "Тепло, Гкал", "Объём, м3", "Газ, Нм3"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(1, col, header)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 16

    row_idx = 2
    for number in sorted(daily):
        for rec in daily[number]:
            ws.cell(row_idx, 1, number).number_format = "0"
            day_cell = ws.cell(row_idx, 2, rec.day)
            day_cell.number_format = "DD.MM.YYYY"
            ws.cell(row_idx, 3, rec.heat_gcal).number_format = "#,##0.000;-#,##0.000;-"
            ws.cell(row_idx, 4, rec.volume_m3).number_format = "#,##0.00;-#,##0.00;-"
            ws.cell(row_idx, 5, rec.gas_nm3).number_format = "#,##0;-#,##0;-"
            for col in range(1, len(headers) + 1):
                ws.cell(row_idx, col).font = Font(name=FONT, size=10)
                ws.cell(row_idx, col).alignment = Alignment(horizontal="center")
            row_idx += 1

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    if row_idx > 2:
        ws.auto_filter.ref = f"A1:E{row_idx - 1}"


# ------------------------------------------------------------ лист примечаний
def _write_notes_sheet(wb, report: ReportData, calorific_kcal: float, threshold: float) -> None:
    ws = wb.create_sheet("Методика и примечания")
    ws.column_dimensions["A"].width = 110
    ws.sheet_view.showGridLines = False

    lines: list[tuple[str, bool]] = [
        ("Методика расчёта", True),
        (
            "Потребление тепла — сумма суточных архивов (параметр Q) по точке учёта "
            "за период, источник: ЛЭРС УЧЁТ, REST API.",
            False,
        ),
        (
            "Расход газа — сумма суточных объёмов газа за тот же период, Нм3. "
            "В ТЗ единица указана как «Нм3/ч» (мгновенный расход); "
            "для расчёта КПД за период нужен объём за период, поэтому в отчёте — Нм3.",
            False,
        ),
        (
            f"КПД, % = Потребление тепла [Гкал] / (Расход газа [Нм3] × "
            f"{calorific_kcal:,.0f} [ккал/Нм3] / 1 000 000) × 100.",
            False,
        ),
        (
            "Коэффициент 10^6 — перевод ккал в Гкал. Без него формула из ТЗ "
            "не даёт процентов (размерности не сходятся).",
            False,
        ),
        (f"Значения КПД ниже {threshold:.0f} % подсвечиваются красным (ячейка B4 — порог).", False),
        (
            "В строке ИТОГО КПД считается только по объектам, где есть учёт газа: "
            "иначе тепло точек без газового узла завышало бы общий результат.",
            False,
        ),
        ("", False),
        ("Источники данных", True),
        ("ЛЭРС УЧЁТ: суточные архивы, опрос по расписанию — 1 раз в неделю.", False),
        ("1С: затраты по объектам, опрос по расписанию — 1 раз в месяц.", False),
        ("", False),
        ("Состояние формирования отчёта", True),
    ]
    for note in report.notes:
        lines.append((note, False))
    lines += [
        ("", False),
        ("Открытые вопросы к заказчику", True),
        (
            "1. Привязка «котельная → потребители тепла → узел учёта газа» и источник "
            "объёма газа (узел учёта газа в ЛЭРС или ввод из 1С/вручную). "
            "Без этой привязки КПД считать не по чему.",
            False,
        ),
        ("2. Теплотворная способность 8200 — низшая или высшая, и по какому документу.", False),
        (
            "3. Соответствие наименований объектов в ЛЭРС и статей затрат в 1С "
            "(нужна таблица соответствий либо общий код объекта).",
            False,
        ),
        (
            "4. Доступ к 1С: адрес базы, учётная запись, имя регистра/отчёта с затратами "
            "по котельным (сейчас работает режим manual из data/costs.xlsx).",
            False,
        ),
    ]

    for idx, (text, is_header) in enumerate(lines, start=1):
        cell = ws.cell(idx, 1, text)
        cell.font = Font(name=FONT, size=11 if is_header else 10, bold=is_header)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def default_filename(start: date, end: date) -> str:
    return f"Отчет_котельные_{start:%Y%m%d}-{end:%Y%m%d}.xlsx"
