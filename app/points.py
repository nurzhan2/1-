"""Чтение перечня точек учёта из файла заказчика (ЛЭРС учет номера точек учета.xlsx)."""
from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook

from .models import MeasurePoint

log = logging.getLogger(__name__)

_HEADER_MARKERS = ("наименование", "номер")


def load_points(path: Path) -> list[MeasurePoint]:
    """Возвращает перечень точек учёта.

    Файл заказчика имеет свободную вёрстку: заголовок таблицы находится не в первой
    строке, данные начинаются со смещением по колонкам. Поэтому ищем строку-заголовок
    по ключевым словам, а дальше читаем всё, где есть текст и целочисленный номер.
    """
    if not Path(path).exists():
        raise FileNotFoundError(f"Не найден перечень точек учёта: {path}")

    wb = load_workbook(path, data_only=True, read_only=True)
    points: list[MeasurePoint] = []
    seen: set[int] = set()

    for ws in wb.worksheets:
        title_col = number_col = None
        for row in ws.iter_rows(values_only=True):
            cells = list(row)

            if title_col is None:
                lowered = [str(c).strip().lower() if c is not None else "" for c in cells]
                if any(_HEADER_MARKERS[0] in c for c in lowered) and any(
                    _HEADER_MARKERS[1] in c for c in lowered
                ):
                    title_col = next(i for i, c in enumerate(lowered) if _HEADER_MARKERS[0] in c)
                    number_col = next(i for i, c in enumerate(lowered) if _HEADER_MARKERS[1] in c)
                continue

            if title_col >= len(cells) or number_col >= len(cells):
                continue
            raw_title, raw_number = cells[title_col], cells[number_col]
            if raw_title is None or raw_number is None:
                continue
            title = str(raw_title).strip()
            try:
                number = int(str(raw_number).strip())
            except ValueError:
                continue
            if not title or number in seen:
                continue
            seen.add(number)
            points.append(MeasurePoint(number=number, title=_clean_title(title)))

    wb.close()
    if not points:
        raise ValueError(f"В файле {path} не найдено ни одной точки учёта")

    log.info("Загружено точек учёта: %d", len(points))
    return sorted(points, key=lambda p: p.number)


def _clean_title(title: str) -> str:
    """Убирает служебные префиксы сортировки («ъ_») и лишние пробелы."""
    title = " ".join(title.split())
    if title.startswith("ъ_"):
        title = title[2:]
    return title
