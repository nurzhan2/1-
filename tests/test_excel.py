"""Сборка xlsx: формула КПД, именованные диапазоны, условное форматирование."""
from __future__ import annotations

from datetime import date

from openpyxl import load_workbook

from app.excel_writer import write_report
from app.models import DailyRecord, MeasurePoint, ReportData, ReportRow
from app.report import build_report


def _sample_report() -> ReportData:
    report = ReportData(period_start=date(2026, 8, 3), period_end=date(2026, 8, 9))
    report.rows = [
        ReportRow(title="Котельная A", point_number=39, heat_gcal=10.0, gas_nm3=1500.0,
                  days_with_data=7, days_expected=7),
        ReportRow(title="Котельная B", point_number=32, heat_gcal=5.0, gas_nm3=900.0,
                  days_with_data=7, days_expected=7),
    ]
    return report


def test_efficiency_is_formula_not_number(tmp_path):
    path = write_report(_sample_report(), tmp_path / "r.xlsx", calorific_kcal=8200, threshold=80)
    wb = load_workbook(path)
    ws = wb["Отчёт"]
    e7 = ws["E7"].value
    assert isinstance(e7, str) and e7.startswith("=")
    # Формула ссылается на именованный диапазон, а не на «зашитое» число
    assert "ТеплотвГаза" in e7
    assert "8200" not in e7


def test_defined_names_present(tmp_path):
    path = write_report(_sample_report(), tmp_path / "r.xlsx", calorific_kcal=8200, threshold=80)
    wb = load_workbook(path)
    assert "ТеплотвГаза" in wb.defined_names
    assert "ПорогКПД" in wb.defined_names
    assert "$B$3" in wb.defined_names["ТеплотвГаза"].attr_text
    assert "$B$4" in wb.defined_names["ПорогКПД"].attr_text
    # В ячейках лежат сами значения-параметры
    ws = wb["Отчёт"]
    assert ws["B3"].value == 8200
    assert ws["B4"].value == 80


def test_conditional_formatting_applied(tmp_path):
    path = write_report(_sample_report(), tmp_path / "r.xlsx", calorific_kcal=8200, threshold=80)
    wb = load_workbook(path)
    ws = wb["Отчёт"]
    ranges = list(ws.conditional_formatting)
    assert ranges, "условное форматирование не навешено"
    # Правило подсветки использует порог из именованного диапазона
    formulas = []
    for cf in ranges:
        for rule in ws.conditional_formatting[cf]:
            formulas.extend(rule.formula)
    assert any("ПорогКПД" in f for f in formulas)


def test_three_sheets_and_open_questions(tmp_path):
    points = [MeasurePoint(39, "Котельная A", lers_id=1001)]
    daily = {39: [DailyRecord(39, date(2026, 8, 3), heat_gcal=1.2, gas_nm3=200.0)]}
    report = build_report(points, daily, {}, date(2026, 8, 3), date(2026, 8, 9))
    path = write_report(report, tmp_path / "r.xlsx", calorific_kcal=8200, threshold=80, daily=daily)
    wb = load_workbook(path)
    assert wb.sheetnames == ["Отчёт", "Суточные данные", "Методика и примечания"]

    notes = wb["Методика и примечания"]
    texts = [notes.cell(r, 1).value for r in range(1, notes.max_row + 1)]
    questions = [t for t in texts if t and t.strip()[:2] in ("1.", "2.", "3.", "4.")]
    # Ровно 4 открытых вопроса к заказчику
    assert len(questions) == 4


def test_no_data_report_still_writes(tmp_path):
    report = ReportData(period_start=date(2026, 8, 3), period_end=date(2026, 8, 9))
    path = write_report(report, tmp_path / "empty.xlsx", calorific_kcal=8200, threshold=80)
    assert path.exists()
